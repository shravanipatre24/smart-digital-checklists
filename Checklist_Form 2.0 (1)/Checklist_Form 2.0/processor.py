import json
import openpyxl
import re
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Alignment
from excel_maps import get_cell_address

# --- CORE SETTINGS ---
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "data" / "checklists"
SUBMISSION_BASE = BASE_DIR / "data" / "submissions"
ACTIVE_DIR = BASE_DIR / "data" / "active"

# Workbook Registry - Must match app.py and excel_maps.py slugs
WORKBOOK_MAP = {
    "fms": "FMS.xlsx",
    "fan_motor_assembly_balancing": "FanMotorAB.xlsx",
    "leak_testing": "LeakTesting.xlsx",
    "module_assembly_testing": "ModuleAssembly.xlsx"
}

def process_submission_to_excel(relative_json_path):
    """
    Processes the operator's JSON submission into the Master Excel.
    relative_json_path: "setup_checklist/leak_testing/B001_(1)_...json"
    """
    
    # 1. IDENTIFY THE SLUG FROM PATH
    path_parts = str(relative_json_path).replace('\\', '/').split('/')
    
    if len(path_parts) < 2:
        print(f"Error: Could not determine slug from path {relative_json_path}")
        return
    
    # The slug is the folder name (e.g., leak_testing)
    checklist_slug = path_parts[1]
    json_filename = path_parts[-1]
    
    # 2. EXTRACTION: Find Batch (B001) and Sequence (1) from Filename
    batch_match = re.search(r"(B\d+)", json_filename)
    seq_match = re.search(r"\((\d+)\)", json_filename)
    
    batch_id = batch_match.group(1) if batch_match else "B001"
    seq = seq_match.group(1) if seq_match else "1"
    
    # 3. LOAD JSON DATA
    json_full_path = SUBMISSION_BASE / relative_json_path
    try:
        with open(json_full_path, 'r') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Error loading JSON at {json_full_path}: {e}")
        return

    print(f"--- Processing: {checklist_slug} | Batch {batch_id} | Sequence {seq} ---")

    # 4. DEFINE WORKBOOK PATHS
    template_name = WORKBOOK_MAP.get(checklist_slug)
    if not template_name:
        print(f"Error: No workbook mapping found for slug: {checklist_slug}")
        return

    template_path = TEMPLATE_DIR / template_name
    active_file = ACTIVE_DIR / f"{batch_id}_{template_name}"

    # 5. LOAD EXCEL (Create Active from Template if it doesn't exist)
    try:
        if not active_file.exists():
            ACTIVE_DIR.mkdir(parents=True, exist_ok=True)
            if not template_path.exists():
                print(f"Error: Template {template_name} not found in {TEMPLATE_DIR}")
                return
            shutil.copy(template_path, active_file)
            print(f"Created new active file: {active_file.name}")

        wb = openpyxl.load_workbook(active_file)
        ws = wb.active 
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return
    
    # 6. WRITE DATA HELPER
    def write_to_excel(data_dict):
        for field_id, value in data_dict.items():
            # Uses the hardcoded mapping in excel_maps.py
            cell = get_cell_address(checklist_slug, str(seq), field_id)
            if cell:
                ws[cell] = str(value)
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # A. Fill Metadata (Date, Shift, etc.)
    write_to_excel(data.get("metadata", {}))

    # B. Fill Sections (The OK marks)
    for section in data.get("sections", []):
        items_dict = {item.get("id"): item.get("value") for item in section.get("items", [])}
        write_to_excel(items_dict)

    # C. Fill Summary (Operator Sign-off)
    write_to_excel(data.get("summary", {}))

    # 7. SAVE
    try:
        wb.save(active_file)
        print(f"--- Success: Saved to {active_file.name} ---")
    except PermissionError:
        print(f"Permission Error: Please close {active_file.name} before saving!")
    
    return str(active_file)

def add_supervisor_approval(batch_id, sequence, approver_name, checklist_slug, template_name):
    """Stamps the supervisor's name and time on the existing active Excel."""
    active_file = ACTIVE_DIR / f"{batch_id}_{template_name}"

    if not active_file.exists():
        return f"Error: {active_file.name} not found."

    try:
        wb = openpyxl.load_workbook(active_file)
        ws = wb.active
        
        approval_time = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Ensure these IDs match your excel_maps.py Row Maps
        approvals = {
            "set_up_approved_by_cc": approver_name,
            "set_up_approved_time": approval_time
        }

        for field_id, value in approvals.items():
            cell = get_cell_address(checklist_slug, str(sequence), field_id)
            if cell:
                ws[cell] = str(value)
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')

        wb.save(active_file)
        return f"Success: {batch_id} Seq({sequence}) approved by {approver_name}"
    
    except PermissionError:
        return "Error: Close the Excel file before approving!"
    except Exception as e:
        return f"Error: {str(e)}"